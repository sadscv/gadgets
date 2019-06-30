#!/usr/bin/env python
# -*- coding:utf-8 _*-
"""
@author: sadscv
@time: 2019/6/21 15:07
@file: class_generator.py
@desc:
"""
import re

import openpyxl as openpyxl


class CONFIG:
    input_path = "docs/江西师范大学2019年本科招生计划一览表（530定）.xlsx"
    output_path = "docs/class_table.xlsx"
    major_list_old_campus = ['音乐学', '音乐表演', '舞蹈学', '表演(健美操方向）',
                       '表演（体育舞蹈方向）', '武术与民族传统体育']


class ClassGenerator:
    def __init__(self):
        self.processed_lines = []
        self.count = 0
        self.config = CONFIG()

    def read_raw(self):
        raw_xls = openpyxl.load_workbook(self.config.input_path)
        print(raw_xls.sheetnames)
        ws = raw_xls.active
        for row in ws.iter_rows(min_row=2, max_row=105):
            major_name = str(row[1].value)
            # test case1：汇总一行为空
            if major_name != '汇总':
                # for each line in excel, loop i time while i is the number of
                # classes of current major
                self.read_line(row, major_name)

    def read_line(self, row, major_name):
        major_population = int(row[13].value)
        for i in range(int(row[16].value)):
            processed_line = {}
            class_population = major_population // (int(row[16].value) - i)
            major_population -= class_population
            tmp = re.match('[\u4e00-\u9fa5]', major_name)
            major_name = major_name[:tmp.endpos]
            tmp1 = re.search('类|S|J', major_name)

            # if class_name contain 类 or S or J, we have to remove it
            if tmp1:
                major_name = major_name[:tmp1.span()[0] + 1]
                if major_name[-1:] != '类':
                    major_name = major_name[:-1]

            # if current major has only one class, we don't need to
            # append class_number after the major name.
            if int(row[16].value) == 1:
                self.count += 1
                class_name = "19级{}班".format(major_name)
            else:
                self.count += 1
                class_name = "19级{}{}班".format(major_name, i + 1)
            processed_line['班级号'] = self.count + 20190000
            processed_line['班级名称'] = class_name
            processed_line['学院'] = row[0].value
            processed_line['班级性质号'] = '01'
            processed_line['年级号'] = '2019/9/1'
            # major_code = self.major_code_in_db(str(row[2].value).lstrip('‘'))
            # processed_line['专业代码'] = major_code if major_code else None
            processed_line['班级人数'] = class_population
            processed_line['学分制状态'] = '1'
            processed_line['专业名称'] = major_name
            if major_name in self.config.major_list_old_campus:
                processed_line['教学区号'] = '01'
            else:
                processed_line['教学区号'] = '02'
            self.processed_lines.append(processed_line)

    def db_init(self):
        database = 'dbo'
        dsn = 'web'
        user = '090134'
        password = 'sadsad'
        odbc_conn_str = "DSN={};UID={};PWD={}".format(dsn, user, password)

        import pyodbc
        print(odbc_conn_str)
        conn = pyodbc.connect(odbc_conn_str)
        return conn, conn.cursor

    def major_code_in_db(self, raw_code):
        conn, cursor = self.db_init()
        if self.have_major_code_in_db(raw_code):
            return raw_code

    def have_major_code_in_db(self, raw_code):
        conn_local, cursor_local = self.db_init()
        sql = "select * " \
              "from 音乐学院小课数据 " \
              "where 教号={};".format(raw_code)
        result = cursor_local.execute(sql).fetchall()
        if result:
            print(result)
            return True
        else:
            return False

    def write2xls(self):
        wb = openpyxl.load_workbook(self.config.output_path)
        ws = wb.active
        # for item in range(len(self.processed_lines)):
        #     print(item)
        index = 0
        for line in self.processed_lines:
            index += 1
            for item in line:
                print(line['班级名称'])
                ws['A' + str(index + 1)] = line['班级号']
                ws['B' + str(index + 1)] = line['班级名称']
                ws['C' + str(index + 1)] = line['班级性质号']
                ws['D' + str(index + 1)] = line['年级号']
                ws['E' + str(index + 1)] = line['班级人数']
                ws['F' + str(index + 1)] = line['学分制状态']
                ws['G' + str(index + 1)] = line['学院']
                # ws['I'+str(index+1)] = line['专业代码']
                ws['J' + str(index + 1)] = line['教学区号']
                ws['K' + str(index + 1)] = line['专业名称']
        wb.save(self.config.output_path)



        # wb.save(self.config.output_path)


if __name__ == '__main__':
    cg = ClassGenerator()
    cg.read_raw()
    cg.write2xls()
